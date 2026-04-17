import pdfplumber
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def run_extraction(pdf_path, output_path, target_day):
    """Scans the PDF and creates the Excel file."""
    if not os.path.exists(pdf_path):
        return f"Error: File '{pdf_path}' not found in the folder."

    all_records = []
    day_str = str(target_day).zfill(2) # Ensures "2" becomes "02"

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table: continue
                
                # Step 1: Find which column matches the Day (01, 02, etc.)
                day_col_idx = -1
                for r_idx in range(min(4, len(table))):
                    for c_idx, cell in enumerate(table[r_idx]):
                        if cell:
                            clean_cell = "".join(cell.split())
                            if clean_cell.startswith(day_str):
                                day_col_idx = c_idx
                                break
                    if day_col_idx != -1: break
                
                if day_col_idx == -1: continue

                # Step 2: Extract rows for teachers
                current_teacher = ""
                for row in table:
                    name_raw = (row[1] or "").replace("\n", " ").strip()
                    period_raw = (row[2] or "").replace("\n", " ").strip()
                    content = (row[day_col_idx] or "").replace("\n", " ").strip()

                    # Save teacher name, ignore header text
                    if name_raw and not any(x in name_raw.upper() for x in ["TÊN", "NGÀY", "CỘNG"]):
                        current_teacher = name_raw

                    if current_teacher and content and content.lower() != "none":
                        slot = None
                        if "1-2" in period_raw: slot = "1 - 2"
                        elif "3-4" in period_raw: slot = "3 - 4"
                        elif "5-6" in period_raw: slot = "5 - 6"
                        elif "7-8" in period_raw: slot = "7 - 8"
                        
                        if slot:
                            all_records.append({
                                "Họ và tên": current_teacher,
                                "môn học": "KB",
                                "slot": slot,
                                "content": content
                            })

        if not all_records:
            return f"No data found in the PDF for Day {day_str}."

        # Step 3: Format and Save to Excel
        df = pd.DataFrame(all_records)
        final_df = df.pivot_table(
            index=["Họ và tên", "môn học"], 
            columns="slot", 
            values="content", 
            aggfunc='first'
        ).reset_index()

        for col in ["1 - 2", "3 - 4", "5 - 6", "7 - 8"]:
            if col not in final_df.columns: final_df[col] = ""
        
        final_df = final_df[["Họ và tên", "môn học", "1 - 2", "3 - 4", "5 - 6", "7 - 8"]]

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, startrow=4)
            ws = writer.sheets['Sheet1']
            ws.merge_cells('A1:F1')
            ws['A1'] = f"KẾ HOẠCH GIẢNG DẠY NGÀY {day_str}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

        return f"Successfully saved to {output_path}"
    except Exception as e:
        return f"Process Error: {str(e)}"