import customtkinter  # This tells Python what 'customtkinter' is
import tkinter
import sys as _sys
for _stream in (_sys.stdout, _sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8")
    except Exception:
        pass
from extractor import run_extraction # Linking to your other file
import customtkinter as ctk
import pandas as pd
from tkinter import messagebox, filedialog
import os
from datetime import datetime
import pdfplumber
import re
from openpyxl.styles import Font, Alignment
from datetime import datetime
today = datetime.now().strftime("%d/%m/%Y")
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
from openpyxl.styles import Alignment, Font, Border
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl.styles import Alignment, Font, Border, Side
import sys
import glob
import customtkinter as ctk
from PIL import Image
import os
import customtkinter as ctk
from PIL import Image
import shutil
from tkinter import filedialog
import pandas as pd
from tkinter import ttk
import subprocess
import subprocess
import os
import subprocess
import platform
def load_documents_view(self, base_directory):
    # CHANGE THIS to your actual scrollable frame variable
    target_frame = self.main_scrollable_frame 

    # Clear existing UI
    for widget in target_frame.winfo_children():
        widget.destroy()

    if not os.path.exists(base_directory):
        print(f"DEBUG: Cannot find path: {base_directory}")
        return

    print(f"DEBUG: Scanning main folder: {base_directory}")

    for item_name in sorted(os.listdir(base_directory)):
        if item_name.startswith('.') or item_name.startswith('~$'):
            continue

        item_path = os.path.join(base_directory, item_name)

        if os.path.isdir(item_path):
            print(f"DEBUG: Found Sub-folder -> {item_name}")
            
            # 1. Create the Folder Frame
            folder_frame = ctk.CTkFrame(target_frame, fg_color="transparent")
            folder_frame.pack(fill="x", pady=5, padx=5)
            ctk.CTkLabel(folder_frame, text=f"📁 {item_name}", font=("Arial", 14, "bold")).pack(anchor="w", padx=5)

            # 2. Scan inside the Sub-folder
            for sub_item in sorted(os.listdir(item_path)):
                if sub_item.startswith('.') or sub_item.startswith('~$'):
                    continue
                    
                sub_item_path = os.path.join(item_path, sub_item)

                if os.path.isfile(sub_item_path):
                    print(f"DEBUG:   Found File -> {sub_item}")
                    
                    # 3. Create File Frame INSIDE Folder Frame
                    file_frame = ctk.CTkFrame(folder_frame)
                    file_frame.pack(fill="x", pady=2, padx=(30, 5)) 
                    ctk.CTkLabel(file_frame, text=f"📄 {sub_item}").pack(side="left", padx=10, pady=5)



def refresh_files(self):
    for widget in self.file_container.winfo_children():
        widget.destroy()

    files = get_all_files("Document")

    for file in files:
        self.create_file_row(file["relative"], file["path"])
def open_file(path):
    if platform.system() == "Darwin":  # macOS
        subprocess.call(["open", path])
    elif platform.system() == "Windows":
        os.startfile(path)
    else:  # Linux
        subprocess.call(["xdg-open", path])
def get_all_files(folder_path):
    all_files = []

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            full_path = os.path.join(root, file)

            # Optional: get relative path for display
            relative_path = os.path.relpath(full_path, folder_path)

            all_files.append({
                "name": file,
                "path": full_path,
                "relative": relative_path
            })

    return all_files

# Hàm quan trọng: Giúp file .exe xác định đúng thư mục đang đứng
def get_base_path():
    if getattr(sys, 'frozen', False):
        # Nếu là file .exe, lấy đường dẫn thư mục chứa file .exe
        return os.path.dirname(sys.executable)
    # Nếu đang chạy code .py trong VS Code
    return os.path.dirname(os.path.abspath(__file__))

def auto_update_schedule():
    # 1. Lấy ngày hiện tại
    today = datetime.now()
    date_str = today.strftime('%Y-%m-%d')
    vn_date = today.strftime('%d/%m/%Y')

    base_path = get_base_path()
    
    # 2. Tự động tìm file có tên chứa "schedule" trong thư mục
    search_pattern = os.path.join(base_path, "schedule*.*")
    files = glob.glob(search_pattern)
    
    if not files:
        print(f"[auto_update_schedule] Không tìm thấy file 'schedule' nào trong: {base_path}")
        return

    file_path = files[0]

    try:
        # 3. Đọc dữ liệu
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, skiprows=3)
        else:
            df = pd.read_excel(file_path, skiprows=3)

        df.columns = [str(col).replace('\n', ' ').strip() for col in df.columns]
        slot_col = [c for c in df.columns if 'Cặp' in c and 'tiết' in c][0]
        
        # Tìm cột ngày hôm nay
        date_column = next((c for c in df.columns if c.startswith(date_str)), None)

        if not date_column:
            print(f"[auto_update_schedule] Hôm nay ({vn_date}) không có lịch dạy trong file nguồn.")
            return

        # 4. Xử lý gộp ô và lọc dữ liệu
        df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill()
        df['MÔN HỌC'] = df['MÔN HỌC'].ffill()

        wb = Workbook()
        ws = wb.active
        
        # Định dạng style
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Tiêu đề và Header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"KẾ HOẠCH GIẢNG DẠY NGÀY {today.day} THÁNG {today.month} NĂM {today.year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align

        headers = ["Họ và tên", "môn học", "1 - 2", "3 - 4", "5 - 6", "7 - 8"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=i, value=h)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = center_align

        # 5. Ghi dữ liệu giáo viên
        current_row = 5
        exclude = ["THỐNG KÊ", "QUÂN SỰ", "QUỐC TẾ", "CÔNG AN", "TỔNG", "CỘNG", "SÁNG", "CHIỀU"]
        
        for teacher in df['HỌ VÀ TÊN'].dropna().unique():
            if any(key in str(teacher).upper() for key in exclude) or len(str(teacher)) < 2:
                continue

            teacher_df = df[df['HỌ VÀ TÊN'] == teacher]
            is_first = True
            
            for subject in teacher_df['MÔN HỌC'].unique():
                if any(key in str(subject).upper() for key in exclude): continue

                sub_df = teacher_df[teacher_df['MÔN HỌC'] == subject]
                slots = {"1 - 2": "", "3 - 4": "", "5 - 6": "", "7 - 8": ""}
                
                for _, row in sub_df.iterrows():
                    s = str(row[slot_col]).strip()
                    if "7 - 9" in s: s = "7 - 8"
                    if s in slots: slots[s] = row[date_column]

                row_vals = [teacher if is_first else "", subject, slots["1 - 2"], slots["3 - 4"], slots["5 - 6"], slots["7 - 8"]]
                for idx, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=current_row, column=idx, value=val)
                    cell.border = thin_border
                    cell.alignment = center_align
                is_first = False
                current_row += 1
            current_row += 1

        # Tự động căn chỉnh cột
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        for col in ['C','D','E','F']: ws.column_dimensions[col].width = 18

        # 6. Lưu file cùng thư mục với EXE
        output_path = os.path.join(base_path, f"KeHoach_Ngay_{date_str}.xlsx")
        wb.save(output_path)
        print(f"[auto_update_schedule] Đã cập nhật lịch ngày {vn_date}. File: {output_path}")

    except Exception as e:
        print(f"[auto_update_schedule] Lỗi hệ thống: {e}")

if __name__ == "__main__":
    auto_update_schedule()
# another
WEEKDAY_MAP = {
    0: "H",
    1: "B",  # Tuesday
    2: "T",  # Wednesday
    3: "N",  # Thursday
    4: "S",  # Friday
    5: "By", # Saturday
    6: "CN"  
}
TIME_WINDOWS = {
    "1-2": ("06:45", "08:15"),
    "3-4": ("08:25", "09:55"),
    "5-6": ("10:05", "11:25"),
    "7-8": ("13:45", "15:05")
}

def check_teaching_status(period_key):
    """Returns '1-2' if teaching, else 'He is out of class'"""
    now = datetime.now().time()
    clean_key = period_key.replace(" ", "") # Handles "1 - 2"
    
    if clean_key in TIME_WINDOWS:
        start_str, end_str = TIME_WINDOWS[clean_key]
        start = datetime.strptime(start_str, "%H:%M").time()
        end = datetime.strptime(end_str, "%H:%M").time()
        
        if start <= now <= end:
            return f"Teaching: {period_key}"
            
    return "He is out of class"

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")
COLORS = {
    "bg": "#F5F7F9",          # Nền chính (Xám nhẹ)
    "sidebar": "#FFFFFF",      # Sidebar (Trắng)
    "card": "#FFFFFF",         # Thẻ nội dung (Trắng)
    "accent": "#2563EB",       # Xanh dương chủ đạo
    "text": "#1E293B",         # Chữ chính (Đen xanh)
    "text_dim": "#64748B",     # Chữ phụ (Xám)
    "hover": "#F1F5F9",        # Màu khi di chuột qua
    "border": "#E2E8F0",       # Màu viền mảnh
    "success": "#10B981",      # Xanh lá (Dùng cho trạng thái sẵn sàng)
    "warning": "#F59E0B",      # Vàng cam
    "error": "#EF4444",        # Đỏ
    "purple": "#8B5CF6",       # Tím (Dự phòng cho các nút cũ)
    "orange": "#F97316",       # Cam (Dự phòng cho các nút cũ)
    "sidebar_dark": "#0F172A", # Sidebar tối (tùy chọn)
}

import json

class AppConfig:
    DEFAULTS = {
        "teacher_file": "danh sách k8.xlsx",
        "schedule_file": "schedule.xlsx",
        "document_folder": "Document",
        "user_name": "Giảng viên",
        "user_email": "",
        "auto_update_schedule": True,
        "appearance": "light",
    }
    PATH = "config.json"

    @classmethod
    def load(cls):
        try:
            with open(cls.PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            merged = dict(cls.DEFAULTS)
            merged.update(data or {})
            return merged
        except Exception:
            return dict(cls.DEFAULTS)

    @classmethod
    def save(cls, data):
        try:
            with open(cls.PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"[config] save error: {e}")
            return False


def clean_numeric_text(val):
    """Chuyển '1.0' -> '1', '2.5' giữ nguyên, NaN/None -> ''."""
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return f"{val:g}"
    s = str(val).strip()
    if s.lower() == "nan":
        return ""
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s


class DocumentFrame(ctk.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        
        # Configure grid
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Create Treeview
        self.tree = ttk.Treeview(self, columns=("filename"), show="headings")
        self.tree.heading("filename", text="Tên tài liệu (Double click to open)")
        self.tree.column("filename", anchor="w", width=400)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        # Bind double click
        self.tree.bind("<Double-1>", self.open_file)

        self.refresh_list()

    def refresh_list(self):
        path = "document"
        if not os.path.exists(path):
            os.makedirs(path)
            
        for file in os.listdir(path):
            if not file.startswith('.'): # Ignore hidden mac files
                self.tree.insert("", "end", values=(file,))

    def open_file(self, event):
        selected_item = self.tree.selection()[0]
        name = self.tree.item(selected_item, "values")[0]
        full_path = os.path.join("document", name)
        
        # Mac specific command to open Word/PDF
        subprocess.call(["open", full_path])
class TeacherCard(ctk.CTkFrame):
    def __init__(self, master, name, period, detail):
        super().__init__(master)
        
        self.period = period
        
        # Teacher Name & Detail (e.g., "td+ b6,7/c2")
        self.info_label = ctk.CTkLabel(self, text=f"{name} ({detail})", font=("Arial", 13))
        self.info_label.pack(side="left", padx=10)
        
        # Status Notification Label
        self.status_label = ctk.CTkLabel(self, text="", font=("Arial", 12, "bold"))
        self.status_label.pack(side="right", padx=10)
     
        self.update_status()  
class TeacherDetailWindow(ctk.CTkToplevel):

    def __init__(self, parent, data):
        super().__init__(parent)
        # Lấy tên giảng viên làm tiêu đề
        name = str(data.get('HỌ VÀ TÊN', 'CHI TIẾT')).upper()
        self.title(f"Thông tin: {name}")
        self.geometry("550x650")
        self.attributes("-topmost", True)  # Luôn hiện trên cùng
        self.configure(fg_color="#F1F5F9")

        # Container chính
        container = ctk.CTkFrame(self, fg_color="white", corner_radius=15)
        container.pack(fill="both", expand=True, padx=20, pady=20)

        ctk.CTkLabel(container, text="HỒ SƠ GIẢNG VIÊN", font=("Arial", 16, "bold"), text_color="#64748B").pack(pady=(15, 0))
        ctk.CTkLabel(container, text=name, font=("Arial", 24, "bold"), text_color="#1E40AF").pack(pady=(0, 20))

        # Vùng cuộn thông tin
        info_scroll = ctk.CTkScrollableFrame(container, fg_color="transparent")
        info_scroll.pack(fill="both", expand=True, padx=10)

        # Tự động quét qua tất cả các cột dữ liệu
        for key, value in data.items():
            if "UNNAMED" in str(key).upper():
                continue
            display_val = clean_numeric_text(value)
            if not display_val:
                continue

            row_frame = ctk.CTkFrame(info_scroll, fg_color="#F8FAFC", corner_radius=8)
            row_frame.pack(fill="x", pady=3)

            ctk.CTkLabel(row_frame, text=str(key), font=("Arial", 12, "bold"),
                         text_color="#475569", width=160, anchor="w"
                         ).pack(side="left", padx=15, pady=10)

            ctk.CTkLabel(row_frame, text=display_val, font=("Arial", 13),
                         text_color="#1E293B", wraplength=280, justify="left"
                         ).pack(side="left", fill="x", expand=True, padx=5)

        ctk.CTkButton(container, text="ĐÓNG", fg_color="#1E293B", command=self.destroy).pack(pady=20)
   
class TeacherManagerPro(ctk.CTk):
    def clear_right_frame(self):
    # Ensure self.right_frame actually exists before trying to clear it
        if hasattr(self, 'right_frame'):
            for widget in self.right_frame.winfo_children():
                widget.destroy()
        else:
            print("Error: right_frame has not been initialized yet.")
    def hide_all_frames(self):
        for name in ("dashboard_frame", "mgmt_frame", "plan_frame",
                     "month_frame", "document_frame", "settings_frame"):
            f = getattr(self, name, None)
            if f is not None:
                f.pack_forget()


    def build_tree(self, folder_path):
        def is_hidden(name):
            return name.startswith(".") or name.startswith("~$") or name.lower() == "thumbs.db"

        tree = {}

        for root, dirs, files in os.walk(folder_path):
            dirs[:] = [d for d in dirs if not is_hidden(d)]

            rel_path = os.path.relpath(root, folder_path)
            parts = rel_path.split(os.sep) if rel_path != "." else []

            current = tree
            for part in parts:
                current = current.setdefault(part, {})

            for file in files:
                if is_hidden(file):
                    continue
                current[file] = None

        return tree
    def render_documents(self):
        if not hasattr(self, "doc_tree"):
            return
        for item in self.doc_tree.get_children():
            self.doc_tree.delete(item)
        self._doc_paths.clear()

        folder = self.config_data.get("document_folder", "Document")
        self.doc_status.configure(text=f"Thư mục: {folder}")

        if not os.path.exists(folder):
            self.doc_tree.insert("", "end",
                text=f"  Không tìm thấy '{folder}'. Mở tab Cài đặt để chọn lại.")
            return

        tree = self.build_tree(folder)
        if not tree:
            self.doc_tree.insert("", "end", text="  Thư mục trống")
            return

        self._populate_doc_tree("", tree, folder)

    def _populate_doc_tree(self, parent, tree, base_path):
        for name, content in sorted(tree.items(),
                                     key=lambda x: (not isinstance(x[1], dict),
                                                    x[0].lower())):
            full = os.path.join(base_path, name)
            if isinstance(content, dict):
                item = self.doc_tree.insert(parent, "end",
                                            text=f"  📁  {name}", open=False)
                self._populate_doc_tree(item, content, full)
            else:
                item = self.doc_tree.insert(parent, "end",
                                            text=f"  📄  {name}")
                self._doc_paths[item] = full

    def _on_doc_tree_activate(self, event=None):
        sel = self.doc_tree.selection()
        if not sel:
            return
        path = self._doc_paths.get(sel[0])
        if not path:
            return
        try:
            if os.name == "nt":
                os.startfile(path)
            else:
                subprocess.call(["open", path])
        except Exception as e:
            self.doc_status.configure(text=f"Lỗi mở file: {e}")
    def render_tree(self, parent, tree, base_path="", level=0):
        for name, content in sorted(tree.items(), key=lambda x: (not isinstance(x[1], dict), x[0].lower())):
            full_path = os.path.join(base_path, name)
            if isinstance(content, dict):
                self._render_folder_node(parent, name, content, full_path, level)
            else:
                self._render_file_node(parent, name, full_path, level)

    def _render_folder_node(self, parent, name, content, full_path, level):
        container = ctk.CTkFrame(parent, fg_color="transparent")
        container.pack(fill="x", padx=0, pady=0, anchor="w")

        header = ctk.CTkFrame(container, fg_color=("#E2E8F0", "#1F2937"), corner_radius=6, height=36)
        header.pack(fill="x", padx=(10 + level * 22, 10), pady=2)
        header.pack_propagate(False)

        child_frame = ctk.CTkFrame(container, fg_color="transparent")
        state = {"open": False}

        arrow = ctk.CTkLabel(header, text="▸", font=("Segoe UI", 12, "bold"), width=18, cursor="hand2")
        arrow.pack(side="left", padx=(10, 0))
        icon = ctk.CTkLabel(header, text="📁", font=("Segoe UI", 13), cursor="hand2")
        icon.pack(side="left", padx=(2, 4))
        label = ctk.CTkLabel(header, text=name, font=("Segoe UI", 13, "bold"), cursor="hand2", anchor="w")
        label.pack(side="left", padx=2, fill="x", expand=True)

        def toggle(event=None):
            if state["open"]:
                child_frame.pack_forget()
                arrow.configure(text="▸")
                icon.configure(text="📁")
                state["open"] = False
            else:
                child_frame.pack(fill="x", anchor="w")
                arrow.configure(text="▾")
                icon.configure(text="📂")
                state["open"] = True

        for w in (header, arrow, icon, label):
            w.bind("<Button-1>", toggle)

        self.render_tree(child_frame, content, full_path, level + 1)

    def _render_file_node(self, parent, name, full_path, level):
        normal_color = ("#F8FAFC", "#111827")
        hover_color = ("#DBEAFE", "#1E3A8A")

        row = ctk.CTkFrame(parent, height=34, corner_radius=6, fg_color=normal_color)
        row.pack(fill="x", padx=(30 + level * 22, 10), pady=2)
        row.pack_propagate(False)

        icon = ctk.CTkLabel(row, text="📄", font=("Segoe UI", 13), cursor="hand2")
        icon.pack(side="left", padx=(10, 4))
        label = ctk.CTkLabel(row, text=name, font=("Segoe UI", 12), cursor="hand2", anchor="w")
        label.pack(side="left", padx=0, fill="x", expand=True)

        def on_click(event=None):
            self.open_document(full_path)

        def on_enter(event=None):
            row.configure(fg_color=hover_color)

        def on_leave(event=None):
            row.configure(fg_color=normal_color)

        for w in (row, icon, label):
            w.bind("<Button-1>", on_click)
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)

    def create_teacher_card(self, row):
        """Hàm phụ tạo từng dòng giảng viên"""
        card = ctk.CTkFrame(self.mgmt_scroll, fg_color="white", height=55, corner_radius=10, 
                            border_width=1, border_color="#E2E8F0")
        card.pack(fill="x", pady=2, padx=(20, 10))
        card.pack_propagate(False)

        # Hiển thị tên
        name_label = ctk.CTkLabel(card, text=row.get('HỌ VÀ TÊN', 'N/A'), font=("Arial", 14, "bold"))
        name_label.pack(side="left", padx=20)
        
        # Hiển thị cấp bậc (nếu có)
        rank = row.get('CẤP BẬC', '')
        if rank and str(rank) != "nan":
            ctk.CTkLabel(card, text=f"({rank})", font=("Arial", 12), text_color="#64748B").pack(side="left")

        # Nút bấm xem chi tiết
        # Truyền toàn bộ dữ liệu của dòng (row) vào cửa sổ mới
        btn = ctk.CTkButton(card, text="XEM CHI TIẾT", width=100, height=32, 
                            fg_color="#2563EB", hover_color="#1D4ED8",
                            command=lambda r=row.to_dict(): TeacherDetailWindow(self, r))
        btn.pack(side="right", padx=15)
   
    def show_document_frame(self):
        self.hide_all_frames()
        self.document_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_document)
        self.render_documents()
  
    def start_live_sync(self):
        if self.mgmt_data and self.plan_path: # Ensure files are linked
            # Update the data in memory
           
            # Re-render the UI
            self.render_mgmt()
            
        # Refresh every 60 seconds to keep "Real Time"
        self.after(60000, self.start_live_sync)
    def process_military_plan_with_calendar(file_path):
        teaching_data = []
        
        # Get current date info
        now = datetime.now()
        today_num = str(now.day).zfill(2) # "31"
        today_char = WEEKDAY_MAP[now.weekday()] # "B" (for Tuesday, March 31)

        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    table = page.extract_table()
                    if not table or len(table) < 2: continue
                    
                    # Row 0 is "Ngày" (01, 02, 03...)
                    # Row 1 is "Thứ" (T, N, S...)
                    days_row = table[0]
                    weekdays_row = table[1]
                    
                    target_col = None
                    for idx in range(len(days_row)):
                        # Check if column matches today's Date AND today's Weekday letter
                        if (days_row[idx] == today_num and 
                            weekdays_row[idx] == today_char):
                            target_col = idx
                            break
                    
                    if target_col is None: continue # Day not found on this page

                    # Process teachers in the rows below
                    for row in table[2:]:
                        # row[2] = Name, row[4] = Period (Tiết)
                        name = " ".join(str(row[2]).split()) if row[2] else None
                        period = str(row[4]).replace(" ", "") if row[4] else None
                        activity = row[target_col] # What is in today's column

                        if name and period and activity and activity.strip():
                            status = check_teaching_status(period)
                            teaching_data.append({
                                "teacher": name,
                                "period": period,
                                "detail": activity.strip(),
                                "notification": status
                            })
        except Exception as e:
            print(f"Error parsing PDF calendar: {e}")
        
        return teaching_data
    def sync_with_military_plan(self):
        # This automatically fetches today's specific assignments from the PDF[cite: 1]
        
        self.render_plan()
        # Check again every 5 minutes to see if a teacher has started a new slot
        self.after(300000, self.sync_with_military_plan)

    
    def __init__(self):
        super().__init__()
        self.title("TSQ Teacher Manager Pro")
        self.geometry("1360x860")
        self.configure(fg_color=COLORS["bg"])

        self.config_data = AppConfig.load()
        ctk.set_appearance_mode(self.config_data.get("appearance", "light"))

        self.mgmt_data = []
        self.plan_data = []

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.setup_sidebar()

        self.main_container = ctk.CTkFrame(self, fg_color="transparent")
        self.main_container.grid(row=0, column=1, sticky="nsew", padx=18, pady=18)

        self.dashboard_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.mgmt_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.plan_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.month_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.document_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")
        self.settings_frame = ctk.CTkFrame(self.main_container, fg_color="transparent")

        self.setup_dashboard_ui()
        self.setup_mgmt_ui()
        self.setup_plan_ui()
        self.setup_month_ui()
        self.setup_document_ui()
        self.setup_settings_ui()

        self.show_dashboard_frame()
        self.update_time()

        self.check_realtime_status()
        self.teacher_db = []
        self.after(100, self.auto_load_mgmt_file)
    
       
    def setup_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=250, corner_radius=0,
                                    fg_color=COLORS["sidebar"],
                                    border_width=0)
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_propagate(False)

        brand = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        brand.pack(fill="x", pady=(22, 16), padx=18)
        logo_badge = ctk.CTkFrame(brand, fg_color=COLORS["accent"], corner_radius=8,
                                  width=32, height=32)
        logo_badge.pack(side="left")
        logo_badge.pack_propagate(False)
        ctk.CTkLabel(logo_badge, text="HD", font=("Arial", 13, "bold"),
                     text_color="white").pack(expand=True)
        ctk.CTkLabel(brand, text="TSQ QLGV", font=("Arial", 17, "bold"),
                     text_color=COLORS["text"]).pack(side="left", padx=10)

        user_card = ctk.CTkFrame(self.sidebar, fg_color=COLORS["bg"], corner_radius=10)
        user_card.pack(fill="x", padx=14, pady=(0, 14))

        avatar = ctk.CTkFrame(user_card, fg_color="#CBD5E1", corner_radius=18,
                              width=36, height=36)
        avatar.pack(side="left", padx=10, pady=8)
        avatar.pack_propagate(False)
        ctk.CTkLabel(avatar, text="GV", font=("Arial", 11, "bold"),
                     text_color="white").pack(expand=True)

        info = ctk.CTkFrame(user_card, fg_color="transparent")
        info.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=8)
        self.lbl_user_name = ctk.CTkLabel(info, text=self.config_data.get("user_name") or "Giảng viên",
                                          font=("Arial", 12, "bold"),
                                          text_color=COLORS["text"], anchor="w")
        self.lbl_user_name.pack(fill="x")
        email = self.config_data.get("user_email") or "Chưa thiết lập"
        self.lbl_user_email = ctk.CTkLabel(info, text=email,
                                           font=("Arial", 10),
                                           text_color=COLORS["text_dim"], anchor="w")
        self.lbl_user_email.pack(fill="x")

        ctk.CTkFrame(self.sidebar, height=1, fg_color=COLORS["border"]
                     ).pack(fill="x", padx=14, pady=(0, 8))

        self.btn_dashboard = self.create_nav_btn("🏠   Bảng điều khiển", self.show_dashboard_frame)
        self.btn_mgmt = self.create_nav_btn("👥   Thông tin giảng viên", self.show_mgmt_frame)
        self.btn_plan = self.create_nav_btn("📅   Kế hoạch ngày", self.show_plan_frame)
        self.btn_month = self.create_nav_btn("🗓   Kế hoạch tháng", self.show_month_frame)
        self.btn_document = self.create_nav_btn("📂   Tài liệu môn học", self.show_document_frame)

        ctk.CTkFrame(self.sidebar, height=1, fg_color=COLORS["border"]
                     ).pack(fill="x", padx=14, pady=(10, 8))

        self.btn_settings = self.create_nav_btn("⚙   Cài đặt", self.show_settings_frame)

        footer = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        footer.pack(side="bottom", fill="x", padx=14, pady=14)
        self.lbl_time = ctk.CTkLabel(footer, text="",
                                     font=("Arial", 11),
                                     text_color=COLORS["text_dim"],
                                     justify="left", anchor="w")
        self.lbl_time.pack(fill="x")
        ctk.CTkLabel(footer, text="v2.0.0 · TSQ Teacher Manager",
                     font=("Arial", 9),
                     text_color=COLORS["text_dim"], anchor="w"
                     ).pack(fill="x", pady=(4, 0))

    def create_nav_btn(self, text, cmd):
        btn = ctk.CTkButton(self.sidebar, text=text, font=("Arial", 13),
                            height=40,
                            fg_color="transparent", text_color=COLORS["text"],
                            anchor="w", hover_color=COLORS["hover"], command=cmd)
        btn.pack(pady=2, padx=12, fill="x")
        return btn

    def set_active_nav(self, active_btn):
        nav_buttons = [
            getattr(self, n, None)
            for n in ("btn_dashboard", "btn_mgmt", "btn_plan", "btn_month",
                      "btn_document", "btn_settings")
        ]
        for b in nav_buttons:
            if b is None:
                continue
            if b is active_btn:
                b.configure(fg_color=COLORS["accent"], text_color="white",
                            hover_color="#1D4ED8",
                            font=("Arial", 13, "bold"))
            else:
                b.configure(fg_color="transparent", text_color=COLORS["text"],
                            hover_color=COLORS["hover"],
                            font=("Arial", 13))
    def load_excel_smart(self, path, check_cols):
        try:
            raw = pd.read_excel(path, header=None)
            header_row = None
            for i, row in raw.iterrows():
                row_vals = [str(x).upper() for x in row.values]
                if any("HỌ VÀ TÊN" in str(val) for val in row_vals):
                    header_row = i
                    break
            
            if header_row is None: return None
            df = pd.read_excel(path, skiprows=header_row)
            df.columns = [str(c).strip() for c in df.columns]
            cols_joined = " ".join(df.columns).upper()
            return df.to_dict('records') if any(col.upper() in cols_joined for col in check_cols) else None
        except: return None
        
    def setup_document_ui(self):
        header = ctk.CTkFrame(self.document_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(header, text="Tài liệu môn học", font=("Arial", 22, "bold"),
                     text_color=COLORS["text"]).pack(side="left")

        ctk.CTkButton(header, text="Làm mới", width=100, height=32,
                      fg_color=COLORS["accent"], hover_color="#1D4ED8",
                      command=self.render_documents).pack(side="right")

        self.doc_status = ctk.CTkLabel(self.document_frame, text="",
                                        font=("Arial", 11),
                                        text_color=COLORS["text_dim"],
                                        anchor="w")
        self.doc_status.pack(fill="x", pady=(0, 6))

        container = ctk.CTkFrame(self.document_frame, fg_color=COLORS["card"],
                                  corner_radius=10, border_width=1,
                                  border_color=COLORS["border"])
        container.pack(fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass
        style.configure("Doc.Treeview", rowheight=30, font=("Segoe UI", 11),
                        background="white", fieldbackground="white",
                        foreground=COLORS["text"], borderwidth=0)
        style.configure("Doc.Treeview.Heading", font=("Segoe UI", 11, "bold"))
        style.map("Doc.Treeview",
                  background=[("selected", COLORS["accent"])],
                  foreground=[("selected", "white")])

        self.doc_tree = ttk.Treeview(container, style="Doc.Treeview",
                                      show="tree", selectmode="browse")
        vsb = ttk.Scrollbar(container, orient="vertical",
                            command=self.doc_tree.yview)
        self.doc_tree.configure(yscrollcommand=vsb.set)
        self.doc_tree.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        vsb.grid(row=0, column=1, sticky="ns")

        self.doc_tree.column("#0", width=700, stretch=True)

        self.doc_tree.bind("<Double-1>", self._on_doc_tree_activate)
        self.doc_tree.bind("<Return>", self._on_doc_tree_activate)

        self._doc_paths = {}

    def setup_month_ui(self):
        header = ctk.CTkFrame(self.month_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(header, text="Kế hoạch tháng", font=("Arial", 22, "bold"),
                     text_color=COLORS["text"]).pack(side="left")

        ctk.CTkButton(header, text="Làm mới", width=100, height=32,
                      fg_color=COLORS["accent"], hover_color="#1D4ED8",
                      command=self.render_month).pack(side="right")

        ctk.CTkButton(header, text="Mở file Excel", width=120, height=32,
                      fg_color="transparent", text_color=COLORS["text"],
                      border_width=1, border_color=COLORS["border"],
                      hover_color=COLORS["hover"],
                      command=self.open_schedule_file).pack(side="right", padx=(0, 8))

        toolbar = ctk.CTkFrame(self.month_frame, fg_color=COLORS["card"],
                                corner_radius=8, border_width=1,
                                border_color=COLORS["border"])
        toolbar.pack(fill="x", pady=(0, 8))

        search_wrap = ctk.CTkFrame(toolbar, fg_color="transparent")
        search_wrap.pack(side="left", fill="y", padx=10, pady=8)
        ctk.CTkLabel(search_wrap, text="🔍", font=("Arial", 12)
                     ).pack(side="left", padx=(0, 4))
        self.month_search = ctk.CTkEntry(search_wrap, height=30, width=200,
                                          placeholder_text="Tìm giảng viên...",
                                          border_color=COLORS["border"])
        self.month_search.pack(side="left")
        self.month_search.bind("<KeyRelease>", lambda e: self.render_month())

        subj_wrap = ctk.CTkFrame(toolbar, fg_color="transparent")
        subj_wrap.pack(side="left", padx=(0, 10), pady=8)
        ctk.CTkLabel(subj_wrap, text="Môn:", font=("Arial", 11),
                     text_color=COLORS["text_dim"]
                     ).pack(side="left", padx=(0, 6))
        self.month_subject = tk.StringVar(value="Tất cả")
        self.month_subject_menu = ctk.CTkOptionMenu(subj_wrap,
                                                     variable=self.month_subject,
                                                     values=["Tất cả"],
                                                     width=110, height=30,
                                                     fg_color=COLORS["accent"],
                                                     button_color=COLORS["accent"],
                                                     button_hover_color="#1D4ED8",
                                                     command=lambda _: self.render_month())
        self.month_subject_menu.pack(side="left")

        self.hide_empty_var = tk.BooleanVar(value=True)
        ctk.CTkSwitch(toolbar, text="Ẩn tiết trống",
                      variable=self.hide_empty_var,
                      font=("Arial", 11),
                      progress_color=COLORS["accent"],
                      command=self.render_month
                      ).pack(side="left", padx=(0, 10), pady=8)

        self.month_info = ctk.CTkLabel(toolbar, text="",
                                        font=("Arial", 11),
                                        text_color=COLORS["text_dim"], anchor="e")
        self.month_info.pack(side="right", padx=10, pady=8)

        container = ctk.CTkFrame(self.month_frame, fg_color=COLORS["card"],
                                 border_width=1, border_color=COLORS["border"],
                                 corner_radius=8)
        container.pack(fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass
        style.configure("Month.Treeview", rowheight=36, font=("Segoe UI", 11),
                        background="white", fieldbackground="white",
                        foreground=COLORS["text"], borderwidth=0)
        style.configure("Month.Treeview.Heading", font=("Segoe UI", 10, "bold"),
                        background="#E0E7FF", foreground=COLORS["text"],
                        padding=(4, 4))
        style.map("Month.Treeview", background=[("selected", COLORS["accent"])],
                  foreground=[("selected", "white")])

        self.month_tree = ttk.Treeview(container, style="Month.Treeview",
                                        show="headings")
        vsb = ttk.Scrollbar(container, orient="vertical",
                             command=self.month_tree.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal",
                             command=self.month_tree.xview)
        self.month_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.month_tree.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self.month_tree.tag_configure("alt", background="#F8FAFC")
        self.month_tree.tag_configure("group_top", background="#EFF6FF",
                                       font=("Segoe UI", 11, "bold"))
        self.month_tree.tag_configure("subj_BC", background="#EFF6FF")
        self.month_tree.tag_configure("subj_DH", background="#F0FDF4")
        self.month_tree.tag_configure("subj_KB", background="#FAF5FF")
        self.month_tree.tag_configure("subj_DN", background="#FFFBEB")

    def show_month_frame(self):
        self.hide_all_frames()
        self.month_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_month)
        self.render_month()

    def open_schedule_file(self):
        path = self.config_data.get("schedule_file", "schedule.xlsx")
        if not os.path.exists(path):
            self.month_info.configure(text=f"Không tìm thấy {path}")
            return
        try:
            if os.name == "nt":
                os.startfile(path)
            else:
                subprocess.call(["open", path])
        except Exception as e:
            self.month_info.configure(text=f"Lỗi mở file: {e}")

    def render_month(self):
        for item in self.month_tree.get_children():
            self.month_tree.delete(item)

        path = self.config_data.get("schedule_file", "schedule.xlsx")
        if not os.path.exists(path):
            self.month_info.configure(
                text=f"Không tìm thấy '{path}'. Vào Cài đặt để chọn lại.")
            self.month_tree.configure(columns=())
            return

        try:
            import re as _re
            raw = pd.read_excel(path, skiprows=3, header=None)
            raw = raw.dropna(how="all").reset_index(drop=True)
            if len(raw) < 2:
                self.month_info.configure(text="File rỗng")
                self.month_tree.configure(columns=())
                return

            header_row = [str(v) if not pd.isna(v) else "" for v in raw.iloc[0]]
            weekday_row = []
            data_start = 1
            if len(raw) > 1:
                second = raw.iloc[1]
                if pd.isna(second.iloc[0]) or str(second.iloc[0]).strip() in ("", "nan"):
                    weekday_row = [str(v) if not pd.isna(v) else "" for v in second]
                    data_start = 2

            data = raw.iloc[data_start:].reset_index(drop=True)
            data.columns = range(len(header_row))
            if data.empty:
                self.month_info.configure(text="Không có dữ liệu")
                return

            for i in range(min(3, len(header_row))):
                data[i] = data[i].ffill()

            search = self.month_search.get().strip().lower() if hasattr(self, "month_search") else ""
            subject_filter = self.month_subject.get() if hasattr(self, "month_subject") else "Tất cả"
            hide_empty = bool(self.hide_empty_var.get()) if hasattr(self, "hide_empty_var") else False

            today = datetime.now()
            today_key = today.strftime("%Y-%m-%d")

            weekday_map_vn = {0: "T2", 1: "T3", 2: "T4", 3: "T5", 4: "T6", 5: "T7", 6: "CN"}

            col_ids = [f"c{i}" for i in range(len(header_row))]
            self.month_tree.configure(columns=col_ids)

            def fmt_date_col(orig):
                m = _re.match(r"(\d{4})-(\d{2})-(\d{2})", str(orig))
                if not m:
                    return str(orig).replace("\n", " ").strip(), False, False
                yyyy, mm, dd = m.group(1), m.group(2), m.group(3)
                date_obj = datetime(int(yyyy), int(mm), int(dd))
                wk_label = weekday_map_vn.get(date_obj.weekday(), "")
                is_weekend = date_obj.weekday() >= 5
                is_today = (today.year == date_obj.year
                            and today.month == date_obj.month
                            and today.day == date_obj.day)
                mark = " ●" if is_today else ""
                label = f"{int(dd):02d}/{int(mm):02d} {wk_label}{mark}"
                return label, is_weekend, is_today

            for i, orig in enumerate(header_row):
                cid = col_ids[i]
                up = str(orig).upper()
                if up == "TT":
                    title = "TT"
                    self.month_tree.column(cid, width=44, minwidth=40,
                                           anchor="center", stretch=False)
                elif "HỌ VÀ TÊN" in up:
                    title = "Họ và tên"
                    self.month_tree.column(cid, width=180, minwidth=140,
                                           anchor="w", stretch=False)
                elif "MÔN" in up:
                    title = "Môn"
                    self.month_tree.column(cid, width=60, minwidth=50,
                                           anchor="center", stretch=False)
                elif _re.match(r"(\d{4})-(\d{2})-(\d{2})", str(orig)):
                    label, is_weekend, is_today = fmt_date_col(orig)
                    title = label
                    width = 110 if is_today else 100
                    self.month_tree.column(cid, width=width, minwidth=80,
                                           anchor="center", stretch=False)
                else:
                    title = str(orig).replace("\n", " ").strip() or " "
                    self.month_tree.column(cid, width=72, minwidth=60,
                                           anchor="center", stretch=False)
                self.month_tree.heading(cid, text=title)

            subjects_seen = set()
            for _, r in data.iterrows():
                s = clean_numeric_text(r[2] if len(r) > 2 else "")
                if s:
                    subjects_seen.add(s.upper())
            if hasattr(self, "month_subject_menu"):
                menu_values = ["Tất cả"] + sorted(subjects_seen)
                try:
                    self.month_subject_menu.configure(values=menu_values)
                except Exception:
                    pass
                if subject_filter not in menu_values:
                    self.month_subject.set("Tất cả")
                    subject_filter = "Tất cả"

            time_slot_col = 3

            teachers = []
            current = None
            for _, r in data.iterrows():
                name = clean_numeric_text(r[1] if len(r) > 1 else "")
                if name and (current is None or current["name"] != name):
                    current = {
                        "tt": clean_numeric_text(r[0]),
                        "name": name,
                        "subject": clean_numeric_text(r[2] if len(r) > 2 else ""),
                        "rows": [],
                    }
                    teachers.append(current)
                if current is None:
                    continue
                current["rows"].append(r)

            subject_tag_map = {
                "BC": "subj_BC", "ĐH": "subj_DH", "DH": "subj_DH",
                "KB": "subj_KB", "ĐN": "subj_DN", "DN": "subj_DN",
            }

            total_rows = 0
            total_teachers = 0
            for t in teachers:
                if search and search not in t["name"].lower():
                    continue
                if subject_filter != "Tất cả" and t["subject"].upper() != subject_filter.upper():
                    continue

                total_teachers += 1
                subj_tag = subject_tag_map.get(t["subject"].upper(), "")

                first_in_block = True
                for r in t["rows"]:
                    values = []
                    for i in range(len(header_row)):
                        v = r[i] if i < len(r) else ""
                        values.append(clean_numeric_text(v).replace("\n", " / "))

                    if hide_empty:
                        has_data = any(values[i].strip() for i in range(4, len(values)))
                        if not has_data:
                            continue

                    if not first_in_block:
                        values[0] = ""
                        values[1] = ""
                        values[2] = ""
                    else:
                        values[0] = t["tt"]
                        values[1] = t["name"]
                        values[2] = t["subject"]

                    tags = []
                    if first_in_block:
                        tags.append("group_top")
                    else:
                        if total_rows % 2 == 1:
                            tags.append("alt")
                    if subj_tag:
                        tags.append(subj_tag)

                    self.month_tree.insert("", "end", values=values, tags=tuple(tags))
                    total_rows += 1
                    first_in_block = False

            hint = " · Hôm nay: " + today.strftime("%d/%m")
            self.month_info.configure(
                text=f"{total_teachers} giảng viên · {total_rows} tiết{hint}")
        except Exception as e:
            self.month_info.configure(text=f"Lỗi đọc file: {e}")
    # --- TAB: QUẢN LÝ CHUNG ---
    def render_mgmt(self):
        for widget in self.mgmt_scroll.winfo_children():
            widget.destroy()

        if not self.mgmt_data:
            ctk.CTkLabel(self.mgmt_scroll, text="Chưa có dữ liệu giảng viên",
                         font=("Arial", 13), text_color=COLORS["text_dim"]).pack(pady=40)
            if hasattr(self, "mgmt_count"):
                self.mgmt_count.configure(text="")
            return

        search = self.mgmt_search.get().lower().strip()

        count = 0
        for row in self.mgmt_data:
            name = str(row.get('HỌ VÀ TÊN', '')).strip()
            if not name or name.lower() == "nan":
                continue
            if search and search not in name.lower():
                continue

            card = ctk.CTkFrame(self.mgmt_scroll, fg_color=COLORS["card"],
                                height=46, corner_radius=8,
                                border_width=1, border_color=COLORS["border"])
            card.pack(fill="x", pady=3, padx=2)
            card.pack_propagate(False)

            ctk.CTkLabel(card, text=name, font=("Arial", 13, "bold"),
                         text_color=COLORS["text"]).pack(side="left", padx=14)

            rank = str(row.get('CẤP BẬC', '')).strip()
            if rank and rank.lower() != "nan":
                ctk.CTkLabel(card, text=rank, font=("Arial", 11),
                             text_color=COLORS["text_dim"]).pack(side="left", padx=(0, 10))

            ctk.CTkButton(card, text="Chi tiết", width=80, height=28,
                          fg_color=COLORS["accent"], hover_color="#1D4ED8",
                          font=("Arial", 11, "bold"),
                          command=lambda r=row: TeacherDetailWindow(self, r)
                          ).pack(side="right", padx=10)
            count += 1

        if hasattr(self, "mgmt_count"):
            total = len(self.mgmt_data)
            self.mgmt_count.configure(text=f"{count}/{total} giảng viên")
    def show_mgmt_frame(self):
        self.hide_all_frames()
        self.mgmt_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_mgmt)

    def setup_dashboard_ui(self):
        header = ctk.CTkFrame(self.dashboard_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 14))
        ctk.CTkLabel(header, text="Bảng điều khiển", font=("Arial", 22, "bold"),
                     text_color=COLORS["text"]).pack(side="left")
        ctk.CTkLabel(header, text=datetime.now().strftime("%A, %d/%m/%Y"),
                     font=("Arial", 12), text_color=COLORS["text_dim"]
                     ).pack(side="right")

        welcome = ctk.CTkFrame(self.dashboard_frame,
                               fg_color=("#EFF6FF", "#1E3A8A"),
                               corner_radius=12, border_width=0)
        welcome.pack(fill="x", pady=(0, 14))
        welcome_pad = ctk.CTkFrame(welcome, fg_color="transparent")
        welcome_pad.pack(fill="x", padx=18, pady=14)
        ctk.CTkLabel(welcome_pad,
                     text=f"Xin chào, {self.config_data.get('user_name') or 'Giảng viên'} 👋",
                     font=("Arial", 16, "bold"),
                     text_color=COLORS["accent"], anchor="w"
                     ).pack(fill="x")
        ctk.CTkLabel(welcome_pad,
                     text="Quản lý thông tin giảng viên, kế hoạch giảng dạy và tài liệu môn học.",
                     font=("Arial", 12),
                     text_color=COLORS["text_dim"], anchor="w"
                     ).pack(fill="x", pady=(2, 0))

        grid = ctk.CTkFrame(self.dashboard_frame, fg_color="transparent")
        grid.pack(fill="x", pady=(0, 14))
        for i in range(4):
            grid.grid_columnconfigure(i, weight=1, uniform="stat")

        self.stat_widgets = {}
        stats = [
            ("teachers", "Tổng giảng viên", "0", "#2563EB", "👥"),
            ("subjects", "Số môn học", "0", "#10B981", "📚"),
            ("today", "Tiết dạy hôm nay", "0", "#F59E0B", "⏰"),
            ("files", "Tài liệu", "0", "#8B5CF6", "📂"),
        ]
        for i, (key, title, value, color, icon) in enumerate(stats):
            card = ctk.CTkFrame(grid, fg_color=COLORS["card"], corner_radius=12,
                                border_width=1, border_color=COLORS["border"])
            card.grid(row=0, column=i, sticky="nsew", padx=6)

            top = ctk.CTkFrame(card, fg_color="transparent")
            top.pack(fill="x", padx=16, pady=(14, 4))
            ctk.CTkLabel(top, text=title, font=("Arial", 11),
                         text_color=COLORS["text_dim"], anchor="w"
                         ).pack(side="left", fill="x", expand=True)
            icon_bg = ctk.CTkFrame(top, fg_color=color, corner_radius=8,
                                   width=32, height=32)
            icon_bg.pack(side="right")
            icon_bg.pack_propagate(False)
            ctk.CTkLabel(icon_bg, text=icon, font=("Arial", 14),
                         text_color="white").pack(expand=True)

            value_lbl = ctk.CTkLabel(card, text=value,
                                     font=("Arial", 26, "bold"),
                                     text_color=COLORS["text"], anchor="w")
            value_lbl.pack(fill="x", padx=16, pady=(0, 14))
            self.stat_widgets[key] = value_lbl

        shortcuts = ctk.CTkFrame(self.dashboard_frame, fg_color=COLORS["card"],
                                 corner_radius=12, border_width=1,
                                 border_color=COLORS["border"])
        shortcuts.pack(fill="both", expand=True)
        ctk.CTkLabel(shortcuts, text="Truy cập nhanh",
                     font=("Arial", 14, "bold"),
                     text_color=COLORS["text"], anchor="w"
                     ).pack(fill="x", padx=18, pady=(14, 8))

        shortcut_grid = ctk.CTkFrame(shortcuts, fg_color="transparent")
        shortcut_grid.pack(fill="x", padx=12, pady=(0, 16))
        for i in range(4):
            shortcut_grid.grid_columnconfigure(i, weight=1, uniform="short")

        shortcut_defs = [
            ("Thông tin giảng viên", "Xem danh sách", self.show_mgmt_frame),
            ("Kế hoạch ngày", "Lịch hôm nay", self.show_plan_frame),
            ("Kế hoạch tháng", "Bảng tháng", self.show_month_frame),
            ("Tài liệu môn", "Mở thư mục", self.show_document_frame),
        ]
        for i, (title, sub, cmd) in enumerate(shortcut_defs):
            btn = ctk.CTkButton(shortcut_grid, text="",
                                fg_color="transparent",
                                hover_color=COLORS["hover"],
                                corner_radius=10, height=70,
                                border_width=1, border_color=COLORS["border"],
                                command=cmd)
            btn.grid(row=0, column=i, sticky="ew", padx=6, pady=4)

            inner = ctk.CTkFrame(btn, fg_color="transparent")
            inner.place(relx=0.5, rely=0.5, anchor="center")
            ctk.CTkLabel(inner, text=title, font=("Arial", 12, "bold"),
                         text_color=COLORS["text"]).pack()
            ctk.CTkLabel(inner, text=sub, font=("Arial", 10),
                         text_color=COLORS["text_dim"]).pack()

    def show_dashboard_frame(self):
        self.hide_all_frames()
        self.dashboard_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_dashboard)
        self.refresh_dashboard_stats()

    def refresh_dashboard_stats(self):
        if not hasattr(self, "stat_widgets"):
            return
        teachers = sum(1 for r in self.mgmt_data if str(r.get('HỌ VÀ TÊN', '')).strip())
        subjects = set()
        for r in self.mgmt_data:
            s = str(r.get('MÔN DẠY', '') or r.get('MÔN HỌC', '') or '').strip()
            if s and s.lower() != 'nan':
                subjects.add(s)
        today_count = 0
        for r in self.plan_data:
            for slot in ("1 - 2", "3 - 4", "5 - 6", "7 - 8"):
                v = str(r.get(slot, "")).strip()
                if v and v.lower() != "nan":
                    today_count += 1
        files = 0
        doc_folder = self.config_data.get("document_folder", "Document")
        if os.path.exists(doc_folder):
            for root, dirs, fs in os.walk(doc_folder):
                dirs[:] = [d for d in dirs if not d.startswith('.')]
                for f in fs:
                    if not f.startswith('.') and not f.startswith('~$'):
                        files += 1

        self.stat_widgets["teachers"].configure(text=str(teachers))
        self.stat_widgets["subjects"].configure(text=str(len(subjects)))
        self.stat_widgets["today"].configure(text=str(today_count))
        self.stat_widgets["files"].configure(text=str(files))

    def setup_settings_ui(self):
        header = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 14))
        ctk.CTkLabel(header, text="Cài đặt", font=("Arial", 22, "bold"),
                     text_color=COLORS["text"]).pack(side="left")
        ctk.CTkLabel(header,
                     text="Cấu hình đường dẫn file và tuỳ chọn hiển thị",
                     font=("Arial", 12),
                     text_color=COLORS["text_dim"]).pack(side="left", padx=(12, 0))

        self.settings_vars = {}

        def add_section(title):
            section = ctk.CTkFrame(self.settings_frame, fg_color=COLORS["card"],
                                    corner_radius=12, border_width=1,
                                    border_color=COLORS["border"])
            section.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(section, text=title, font=("Arial", 14, "bold"),
                         text_color=COLORS["text"], anchor="w"
                         ).pack(fill="x", padx=18, pady=(14, 8))
            return section

        def add_file_row(parent, label, key, mode="file",
                         filetypes=(("Excel", "*.xlsx *.xls"),)):
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", padx=18, pady=(0, 10))
            ctk.CTkLabel(row, text=label, font=("Arial", 12),
                         text_color=COLORS["text"], width=170, anchor="w"
                         ).pack(side="left")
            var = tk.StringVar(value=self.config_data.get(key, ""))
            self.settings_vars[key] = var
            entry = ctk.CTkEntry(row, textvariable=var, height=34,
                                 border_color=COLORS["border"])
            entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

            def pick():
                if mode == "folder":
                    p = filedialog.askdirectory(initialdir=".")
                else:
                    p = filedialog.askopenfilename(filetypes=filetypes,
                                                   initialdir=".")
                if p:
                    rel = os.path.relpath(p, os.getcwd())
                    if not rel.startswith(".."):
                        p = rel
                    var.set(p)

            ctk.CTkButton(row, text="Chọn...", width=84, height=34,
                          fg_color="transparent",
                          text_color=COLORS["text"],
                          border_width=1, border_color=COLORS["border"],
                          hover_color=COLORS["hover"], command=pick
                          ).pack(side="left")

        files_section = add_section("Đường dẫn dữ liệu")
        add_file_row(files_section, "File danh sách GV", "teacher_file",
                     mode="file", filetypes=(("Excel", "*.xlsx *.xls"),))
        add_file_row(files_section, "File kế hoạch tháng", "schedule_file",
                     mode="file", filetypes=(("Excel", "*.xlsx *.xls"),))
        add_file_row(files_section, "Thư mục tài liệu", "document_folder",
                     mode="folder")

        user_section = add_section("Thông tin người dùng")
        for label, key in (("Tên hiển thị", "user_name"), ("Email", "user_email")):
            row = ctk.CTkFrame(user_section, fg_color="transparent")
            row.pack(fill="x", padx=18, pady=(0, 10))
            ctk.CTkLabel(row, text=label, font=("Arial", 12),
                         text_color=COLORS["text"], width=170, anchor="w"
                         ).pack(side="left")
            var = tk.StringVar(value=self.config_data.get(key, ""))
            self.settings_vars[key] = var
            ctk.CTkEntry(row, textvariable=var, height=34,
                         border_color=COLORS["border"]
                         ).pack(side="left", fill="x", expand=True)

        pref_section = add_section("Hiển thị")
        row = ctk.CTkFrame(pref_section, fg_color="transparent")
        row.pack(fill="x", padx=18, pady=(0, 10))
        ctk.CTkLabel(row, text="Chế độ giao diện", font=("Arial", 12),
                     text_color=COLORS["text"], width=170, anchor="w"
                     ).pack(side="left")
        appear_var = tk.StringVar(value=self.config_data.get("appearance", "light"))
        self.settings_vars["appearance"] = appear_var
        appear_menu = ctk.CTkOptionMenu(row, variable=appear_var,
                                        values=["light", "dark", "system"],
                                        fg_color=COLORS["accent"],
                                        button_color=COLORS["accent"],
                                        button_hover_color="#1D4ED8",
                                        width=140)
        appear_menu.pack(side="left")

        row2 = ctk.CTkFrame(pref_section, fg_color="transparent")
        row2.pack(fill="x", padx=18, pady=(0, 14))
        ctk.CTkLabel(row2, text="Tự cập nhật lịch ngày",
                     font=("Arial", 12),
                     text_color=COLORS["text"], width=170, anchor="w"
                     ).pack(side="left")
        auto_var = tk.BooleanVar(value=bool(self.config_data.get("auto_update_schedule", True)))
        self.settings_vars["auto_update_schedule"] = auto_var
        ctk.CTkSwitch(row2, text="", variable=auto_var,
                      progress_color=COLORS["accent"]).pack(side="left")

        actions = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        actions.pack(fill="x", pady=(4, 0))
        self.settings_status = ctk.CTkLabel(actions, text="",
                                            font=("Arial", 11),
                                            text_color=COLORS["text_dim"])
        self.settings_status.pack(side="left")
        ctk.CTkButton(actions, text="Lưu cài đặt", width=130, height=36,
                      fg_color=COLORS["accent"], hover_color="#1D4ED8",
                      font=("Arial", 12, "bold"),
                      command=self.save_settings).pack(side="right")
        ctk.CTkButton(actions, text="Tải lại dữ liệu", width=130, height=36,
                      fg_color="transparent", text_color=COLORS["text"],
                      border_width=1, border_color=COLORS["border"],
                      hover_color=COLORS["hover"],
                      command=self.reload_all_data).pack(side="right", padx=(0, 8))

    def show_settings_frame(self):
        self.hide_all_frames()
        self.settings_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_settings)

    def save_settings(self):
        for key, var in self.settings_vars.items():
            try:
                self.config_data[key] = var.get()
            except Exception:
                pass
        ok = AppConfig.save(self.config_data)
        if ok:
            self.settings_status.configure(
                text=f"Đã lưu · {datetime.now().strftime('%H:%M:%S')}",
                text_color=COLORS["success"])
            self.lbl_user_name.configure(text=self.config_data.get("user_name") or "Giảng viên")
            self.lbl_user_email.configure(text=self.config_data.get("user_email") or "Chưa thiết lập")
            try:
                ctk.set_appearance_mode(self.config_data.get("appearance", "light"))
            except Exception:
                pass
        else:
            self.settings_status.configure(text="Lỗi khi lưu config.json",
                                           text_color=COLORS["error"])

    def reload_all_data(self):
        self.auto_load_mgmt_file()
        self.check_realtime_status()
        if hasattr(self, "month_tree"):
            self.render_month()
        if hasattr(self, "document_scroll"):
            self.render_documents()
        self.refresh_dashboard_stats()
        self.settings_status.configure(
            text=f"Đã tải lại · {datetime.now().strftime('%H:%M:%S')}",
            text_color=COLORS["success"])

    def setup_mgmt_ui(self):
        self.clear_right_frame()
        header = ctk.CTkFrame(self.mgmt_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(header, text="Thông tin giảng viên", font=("Arial", 22, "bold"),
                     text_color=COLORS["text"]).pack(side="left")
        self.mgmt_count = ctk.CTkLabel(header, text="", font=("Arial", 12),
                                       text_color=COLORS["text_dim"])
        self.mgmt_count.pack(side="right")

        self.mgmt_search = ctk.CTkEntry(self.mgmt_frame,
                                        placeholder_text="Tìm theo tên giảng viên...",
                                        height=36, border_color=COLORS["border"])
        self.mgmt_search.pack(fill="x", pady=(0, 10))
        self.mgmt_search.bind("<KeyRelease>", lambda e: self.render_mgmt())

        self.mgmt_scroll = ctk.CTkScrollableFrame(self.mgmt_frame, fg_color="transparent")
        self.mgmt_scroll.pack(fill="both", expand=True)
    def link_mgmt(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self.mgmt_path = path
            self.refresh_mgmt()
    def refresh_mgmt(self):

        path = self.config_data.get("teacher_file", "danh sách k8.xlsx")
        if not os.path.exists(path):
            print(f"Không tìm thấy file {path}!")
            return

        try:
            # 1. Đọc file với engine openpyxl (quan trọng)
            # skiprows=2: Bỏ qua các dòng tiêu đề rỗng phía trên
            df = pd.read_excel(path, skiprows=2, engine='openpyxl', dtype=str)

            # 2. Chuẩn hóa tên cột: Xóa khoảng trắng và viết HOA toàn bộ
            df.columns = [str(c).strip().upper() for c in df.columns]

            # 3. Làm sạch dữ liệu: Xử lý gộp ô (ffill) và xóa dòng trống
            if 'HỌ VÀ TÊN' in df.columns:
                df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill() # Điền tên cho các ô bị gộp
                df = df.dropna(subset=['HỌ VÀ TÊN']) # Xóa dòng rác

            # 4. QUAN TRỌNG: Lưu vào biến self để các hàm khác có thể dùng
            self.mgmt_data = df.to_dict('records')
            
            # 5. Sau khi nhận được dữ liệu, gọi hàm vẽ giao diện ngay
            self.render_mgmt()
            print(f"Đã nhận {len(self.mgmt_data)} giảng viên từ file.")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể nhận dữ liệu từ file: {e}")
    def auto_load_mgmt_file(self):
        """Tự động tìm tiêu đề và nạp dữ liệu chính xác"""
        file_name = self.config_data.get("teacher_file", "danh sách k8.xlsx")
        if not os.path.exists(file_name):
            print(f"Không tìm thấy file tại: {os.path.abspath(file_name)}")
            return

        try:
            # 1. Đọc nháp toàn bộ file (không dùng header) để tìm dòng tiêu đề
            raw_df = pd.read_excel(file_name, header=None, engine='openpyxl')
            
            header_row_index = None
            # Quét qua 20 dòng đầu tiên để tìm chữ "HỌ VÀ TÊN"
            for i, row in raw_df.head(20).iterrows():
                # Chuyển tất cả giá trị trong dòng thành chữ HOA, xóa khoảng trắng để so sánh
                row_values = [str(val).strip().upper() for val in row.values]
                if "HỌ VÀ TÊN" in row_values:
                    header_row_index = i
                    print(f"🎯 Đã tìm thấy tiêu đề 'HỌ VÀ TÊN' tại dòng thứ: {i + 1}")
                    break
            
            if header_row_index is None:
                print("❌ Vẫn không tìm thấy cột 'HỌ VÀ TÊN'.")
                print(f"Dữ liệu 5 dòng đầu đọc được:\n{raw_df.head(5)}")
                return

            # 2. Đọc lại file thật sự bắt đầu từ dòng tiêu đề đã tìm thấy
            df = pd.read_excel(file_name, skiprows=header_row_index,
                               engine='openpyxl', dtype=str)

            # 3. Chuẩn hóa tên cột một lần nữa cho chắc chắn
            df.columns = [str(c).strip().upper() for c in df.columns]
            
            # 4. Làm sạch dữ liệu
            # Loại bỏ các cột "Unnamed" (cột thừa không có tên)
            df = df.loc[:, ~df.columns.str.contains('^UNNAMED')]
            
            # Điền đầy dữ liệu nếu có gộp ô (Merge Cells)
            if 'HỌ VÀ TÊN' in df.columns:
                df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill()
                df = df.dropna(subset=['HỌ VÀ TÊN']) # Xóa dòng hoàn toàn trống
                
                # Chuyển đổi sang danh sách Dictionary để dùng cho app
                self.mgmt_data = df.to_dict('records')
                
                # 5. Cập nhật giao diện
                self.render_mgmt()
                print(f"✅ Nạp thành công {len(self.mgmt_data)} giảng viên.")
            
        except Exception as e:
            print(f"❌ Lỗi xử lý: {e}")
    def process_mgmt_file(self, path):
        try:
            # 1. Đọc toàn bộ file không bỏ qua dòng nào để dò tìm
            raw_df = pd.read_excel(path, header=None, engine='openpyxl')
            
            header_row_index = None
            
            # 2. Vòng lặp tìm dòng chứa từ khóa "HỌ VÀ TÊN"
            for i, row in raw_df.iterrows():
                # Chuyển dòng thành danh sách chữ HOA để so sánh
                row_values = [str(val).strip().upper() for val in row.values]
                if "HỌ VÀ TÊN" in row_values:
                    header_row_index = i
                    break
            
            if header_row_index is None:
                print(f"❌ Không tìm thấy dòng nào chứa cột 'HỌ VÀ TÊN' trong file {path}")
                return

            # 3. Đọc lại file với đúng dòng tiêu đề đã tìm thấy
            df = pd.read_excel(path, skiprows=header_row_index, engine='openpyxl')
            
            # 4. Chuẩn hóa tên cột (Xóa khoảng trắng, viết HOA)
            df.columns = [str(c).strip().upper() for c in df.columns]
            
            # 5. Làm sạch dữ liệu rác
            # Điền đầy dữ liệu gộp ô (Merge cells)
            if 'HỌ VÀ TÊN' in df.columns:
                df['HỌ VÀ TÊN'] = df['HỌ VÀ TÊN'].ffill()
                df = df.dropna(subset=['HỌ VÀ TÊN']) # Bỏ dòng trống hoàn toàn
                
                # Chuyển thành List Dict để dùng cho App
                self.mgmt_data = df.to_dict('records')
                
                # Vẽ lên màn hình
                self.render_mgmt()
                print(f"✅ Đã tìm thấy tiêu đề ở dòng {header_row_index + 1} và nạp thành công!")
            else:
                print("❌ Lỗi logic: Đã tìm thấy dòng tiêu đề nhưng không khớp cột.")

        except Exception as e:
            print(f"❌ Lỗi xử lý file: {e}")
   
#--------------------------
    def show_plan_frame(self):
        self.hide_all_frames()
        self.plan_frame.pack(fill="both", expand=True)
        self.set_active_nav(self.btn_plan)

    def render_plan(self):
        try:
            if not self.plan_scroll.winfo_exists():
                return
            for child in self.plan_scroll.winfo_children():
                child.destroy()

            if not self.plan_data:
                ctk.CTkLabel(self.plan_scroll, text="Chưa có dữ liệu kế hoạch ngày",
                             font=("Arial", 13),
                             text_color=COLORS["text_dim"]).pack(pady=40)
                return

            SUB_COLORS = {
                "BC": ("#E0F2FE", "#0369A1"), "ĐH": ("#DCFCE7", "#15803D"),
                "KB": ("#F3E8FF", "#7E22CE"), "ĐN": ("#FEF3C7", "#B45309"),
            }

            header_f = ctk.CTkFrame(self.plan_scroll, fg_color="#F8FAFC", height=36, corner_radius=0)
            header_f.pack(fill="x", pady=(0, 2))
            header_f.pack_propagate(False)

            COLS = [("Họ và tên", 0.02, 0.28), ("Môn", 0.30, 0.10), ("1-2", 0.42, 0.14),
                    ("3-4", 0.57, 0.14), ("5-6", 0.72, 0.14), ("7-8", 0.87, 0.14)]
            for txt, rx, rw in COLS:
                ctk.CTkLabel(header_f, text=txt, font=("Arial", 11, "bold"),
                             text_color=COLORS["text_dim"], anchor="w"
                             ).place(relx=rx, rely=0.5, anchor="w", relwidth=rw)

            prev_name = ""
            group_frame = None
            rendered = 0

            for i, row in enumerate(self.plan_data):
                def get_v(k, _r=row):
                    v = str(_r.get(k, "")).strip()
                    return "" if v.lower() == "nan" or v == "" else v

                full_name = get_v("Họ và tên")
                subject = get_v("môn học")
                if not full_name and not subject:
                    continue

                if full_name and full_name != prev_name:
                    group_frame = ctk.CTkFrame(self.plan_scroll,
                                               fg_color="white", corner_radius=0)
                    group_frame.pack(fill="x", pady=(2, 0))
                    prev_name = full_name
                    is_duplicate = False
                else:
                    is_duplicate = True

                row_f = ctk.CTkFrame(group_frame, fg_color="transparent",
                                     height=32, corner_radius=0)
                row_f.pack(fill="x")
                row_f.pack_propagate(False)

                display_name = full_name if not is_duplicate else ""
                ctk.CTkLabel(row_f, text=display_name, font=("Arial", 12, "bold"),
                             text_color=COLORS["text"], anchor="w"
                             ).place(relx=0.02, rely=0.5, anchor="w", relwidth=0.28)

                if subject:
                    bg, fg = SUB_COLORS.get(subject.upper(), ("#F1F5F9", "#475569"))
                    badge = ctk.CTkFrame(row_f, fg_color=bg, corner_radius=4, height=20)
                    badge.place(relx=0.30, rely=0.5, anchor="w", relwidth=0.08)
                    ctk.CTkLabel(badge, text=subject, font=("Arial", 10, "bold"),
                                 text_color=fg).pack(expand=True)

                for idx, t_col in enumerate(["1 - 2", "3 - 4", "5 - 6", "7 - 8"]):
                    val = get_v(t_col)
                    if val:
                        ctk.CTkLabel(row_f, text=val, font=("Arial", 11),
                                     text_color=COLORS["accent"], anchor="w"
                                     ).place(relx=0.42 + (idx * 0.15), rely=0.5,
                                             anchor="w", relwidth=0.14)
                rendered += 1

            if rendered == 0:
                ctk.CTkLabel(self.plan_scroll, text="Không có dòng nào hợp lệ",
                             font=("Arial", 12),
                             text_color=COLORS["text_dim"]).pack(pady=20)

            self.update_idletasks()
        except Exception as e:
            print(f"Lỗi render_plan: {e}")
    def setup_plan_ui(self):
        self.clear_right_frame()
        header = ctk.CTkFrame(self.plan_frame, fg_color="transparent")
        header.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(header, text="Kế hoạch giảng dạy trong ngày",
                     font=("Arial", 20, "bold"), text_color=COLORS["text"]).pack(side="left")

        ctk.CTkButton(header, text="Làm mới", width=100, height=32,
                      fg_color=COLORS["accent"], hover_color="#1D4ED8",
                      font=("Arial", 11, "bold"),
                      command=self.check_realtime_status).pack(side="right")

        self.status_indicator = ctk.CTkLabel(header, text="● Sẵn sàng",
                                             font=("Arial", 11),
                                             text_color=COLORS["success"])
        self.status_indicator.pack(side="right", padx=(0, 12))

        self.plan_scroll = ctk.CTkScrollableFrame(self.plan_frame,
                                                  fg_color=COLORS["card"],
                                                  border_width=1,
                                                  border_color=COLORS["border"])
        self.plan_scroll.pack(fill="both", expand=True)
    def link_plan(self):
        """Hàm chọn file Excel từ máy tính"""
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.plan_path = path
            self.refresh_plan_data()
    def refresh_plan_data(self):
        if not self.plan_path:
            return

        try:
            df = pd.read_excel(self.plan_path, skiprows=3)
            df.columns = [str(col).strip() for col in df.columns]

            # Fill merged cells
            df['Họ và tên'] = df['Họ và tên'].ffill()

            slots = ["1 - 2", "3 - 4", "5 - 6", "7 - 8"]
            df = df.dropna(subset=['môn học'] + slots, how='all')

            # 🔥 GROUP BY TEACHER
            grouped = []

            for name, group in df.groupby('Họ và tên', sort=False):
                teacher = {
                    "name": name,
                    "subjects": [],
                    "rows": group.to_dict('records')
                }

                for _, r in group.iterrows():
                    subject = str(r.get('môn học', ''))
                    if subject != "nan" and subject not in teacher["subjects"]:
                        teacher["subjects"].append(subject)

                grouped.append(teacher)

            self.plan_data = df.to_dict('records')
            self.render_plan()

        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
    def open_document(self, file_name):
        path = os.path.join("Document", file_name)
        try:
            if os.name == "nt":
                os.startfile(path)
            else:
                subprocess.call(["open", path])
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
    def convert_excel_date(self, val):
        """Hàm phụ trợ để xử lý ngày tháng từ số Excel sang chuỗi dd/mm/yyyy"""
        try:
            if isinstance(val, (int, float)) and val > 1000:
                return pd.to_datetime(val, unit='D', origin='1899-12-30').strftime('%d/%m/%Y')
            return str(val) if str(val).lower() != 'nan' else ""
        except:
            return str(val)
    def update_time(self):
        self.lbl_time.configure(text=datetime.now().strftime("%H:%M:%S\n%A, %d/%m/%Y"))
        self.after(1000, self.update_time)   
    def load_monthly_plan(self, file_path):
        try:
            df = pd.read_excel(file_path)

            # Clear old content if reload
            for widget in self.tab_plan.winfo_children():
                widget.destroy()

            # Frame container
            frame = ctk.CTkFrame(self.tab_plan)
            frame.pack(fill="both", expand=True)

            # Create table
            tree = ttk.Treeview(frame)
            tree.pack(side="left", fill="both", expand=True)

            # Scrollbars
            scrollbar_y = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            scrollbar_y.pack(side="right", fill="y")

            scrollbar_x = ttk.Scrollbar(self.tab_plan, orient="horizontal", command=tree.xview)
            scrollbar_x.pack(fill="x")

            tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

            # Columns
            tree["columns"] = list(df.columns)
            tree["show"] = "headings"

            for col in df.columns:
                tree.heading(col, text=col)
                tree.column(col, anchor="center", width=120)

            # Rows
            for _, row in df.iterrows():
                tree.insert("", "end", values=list(row))

        except Exception as e:
            print("ERROR loading Excel:", e)   
    def load_documents(self, folder_path):
        frame = ctk.CTkFrame(self.tab_docs)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        files = [f for f in os.listdir(folder_path) if f.endswith(".pdf")]

        for file in files:
            btn = ctk.CTkButton(
                frame,
                text=file,
                anchor="w",
                command=lambda f=file: self.open_pdf(os.path.join(folder_path, f))
            )
            btn.pack(fill="x", pady=5)
    def check_realtime_status(self):
        import glob
        files = glob.glob("KeHoach_Ngay_*.xlsx")
        if not files: 
            print("❌ Không tìm thấy file Excel nào!")
            return

        latest_file = max(files, key=os.path.getctime)
        print(f"📂 Đang đọc file: {latest_file}")
        
        try:
            # Đọc từ dòng 4
            df = pd.read_excel(latest_file, skiprows=3)
            df.columns = [str(c).strip() for c in df.columns]
            
            # XỬ LÝ QUAN TRỌNG: Loại bỏ các dòng hoàn toàn trống
            # Chỉ giữ lại dòng có tên HOẶC có môn học
            df = df.dropna(subset=['Họ và tên', 'môn học'], how='all')
            
            self.plan_data = df.to_dict('records')
            
            # KIỂM TRA: In ra số lượng dòng Python đọc được
            print(f"✅ Đã nạp được {len(self.plan_data)} dòng dữ liệu.")
            
            if self.plan_scroll.winfo_exists():
                self.render_plan()
        except Exception as e:
            print(f"❌ Lỗi đọc file: {e}")


if __name__ == "__main__":
    app = TeacherManagerPro()
    app.mainloop()